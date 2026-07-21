import hashlib
import io
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[2]
IMAGE_XLSX = Path("/Users/congjiang/Downloads/商品图x商品名对照表.xlsx")
CATEGORY_XLSX = Path("/Users/congjiang/Downloads/拆商品类目划分赛道明细.xlsx")
OUT_DIR = ROOT / "site/assets/products"
OUT_JS = ROOT / "site/data/product-assets.js"
SELECTION_JS = ROOT / "site/data/selection.js"

TRACK_NAMES = {
    ("生活日用", "功效品"): "生活日用-功效品",
    ("生活日用", "清洁工具"): "生活日用-清洁工具",
    ("生活日用", "收纳用品"): "生活日用-收纳用品",
    ("生活日用", "其他"): "生活日用-其他",
    ("家居家纺", "家居家纺"): "家居家纺-家纺",
    ("家居家纺", "家居工艺品"): "家居家纺-家居工艺品",
    ("厨具", "厨具"): "餐厨水具-厨具",
    ("厨具", "餐具水具"): "餐厨水具-餐具水具",
}


def text(value):
    return str(value or "").strip()


def norm(value):
    value = text(value).lower()
    value = re.sub(r"plant\s*master|植物大师|春禾", "", value, flags=re.I)
    value = re.sub(r"[\s·|｜（）()【】\[\]{}，,。:：;；'\"_-]+", "", value)
    return value.replace("/", "")


def load_categories():
    sheet = load_workbook(CATEGORY_XLSX, data_only=True)["最新表"]
    path_tracks = {}
    leaf_tracks = {}
    parent = ""
    for row in range(2, sheet.max_row + 1):
        parent = text(sheet.cell(row, 1).value) or parent
        child = text(sheet.cell(row, 2).value)
        track = TRACK_NAMES.get((parent, child))
        categories = text(sheet.cell(row, 3).value)
        if not track or not categories or categories.startswith("其他所有"):
            continue
        for category in categories.splitlines():
            category = text(category).replace(" > ", "-")
            if not category:
                continue
            path_tracks[norm(category)] = track
            leaf = norm(category.split("-")[-1])
            if leaf and leaf not in leaf_tracks:
                leaf_tracks[leaf] = track
    return path_tracks, leaf_tracks


def classify(category, leaf, path_tracks, leaf_tracks):
    category_key = norm(text(category).replace(" > ", "-"))
    if category_key in path_tracks:
        return path_tracks[category_key]
    leaf_key = norm(leaf or text(category).replace(" > ", "-").split("-")[-1])
    return leaf_tracks.get(leaf_key, "")


def load_selection():
    raw = SELECTION_JS.read_text(encoding="utf-8")
    match = re.search(r"window\.SELECTION_DATA\s*=\s*(\{[\s\S]*\});", raw)
    if not match:
        raise RuntimeError("selection.js 数据格式无法识别")
    return json.loads(match.group(1))


def all_items(data):
    items = []
    items.extend(data["nonClosed"]["items"])
    items.extend(data["closed"]["channels"]["quanyutong"]["items"])
    items.extend(data["closed"]["channels"]["adq"]["items"])
    for cycle in data.get("cycles", {}).values():
        items.extend(cycle.get("items", []))
    return items


def best_name(name, known):
    if name in known:
        return name
    key = norm(name)
    exact = [candidate for candidate in known if norm(candidate) == key]
    if exact:
        return exact[0]
    parts = [norm(part) for part in re.split(r"[/／]", name) if norm(part)]
    ranked = []
    for candidate in known:
        candidate_key = norm(candidate)
        score = 0
        if candidate_key and (candidate_key in key or key in candidate_key):
            score = min(len(candidate_key), len(key))
        for part in parts:
            if candidate_key == part:
                score = max(score, len(part) + 20)
            elif len(part) >= 2 and (part in candidate_key or candidate_key in part):
                score = max(score, min(len(part), len(candidate_key)))
        if score >= 2:
            ranked.append((score, -abs(len(candidate_key) - len(key)), candidate))
    ranked.sort(reverse=True)
    return ranked[0][2] if ranked else ""


def main():
    path_tracks, leaf_tracks = load_categories()
    data = load_selection()
    items = all_items(data)
    wanted = {text(item.get("name")) for item in items if text(item.get("name"))}

    workbook = load_workbook(IMAGE_XLSX, data_only=True)
    sheet = workbook.active
    rows = {}
    for row in range(2, sheet.max_row + 1):
        name = text(sheet.cell(row, 2).value)
        if name and name not in rows:
            rows[name] = {
                "row": row,
                "category": text(sheet.cell(row, 5).value),
            }

    matched = {}
    for name in wanted:
        source = best_name(name, rows)
        if source:
            matched[name] = source

    row_images = {}
    for image in sheet._images:
        row = image.anchor._from.row + 1
        if row not in row_images:
            row_images[row] = image

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    image_paths = {}
    extracted_sources = {}
    for source, info in sorted(rows.items()):
        image = row_images.get(info["row"])
        if not image:
            continue
        filename = hashlib.sha1(source.encode("utf-8")).hexdigest()[:16] + ".webp"
        destination = OUT_DIR / filename
        picture = Image.open(io.BytesIO(image._data())).convert("RGB")
        picture.thumbnail((480, 480), Image.Resampling.LANCZOS)
        picture.save(destination, "WEBP", quality=76, method=6)
        extracted_sources[source] = f"assets/products/{filename}"
        image_paths[source] = extracted_sources[source]
    for target, source in matched.items():
        if source in extracted_sources:
            image_paths[target] = extracted_sources[source]

    product_tracks = {}
    for source, info in rows.items():
        track = classify(info["category"], "", path_tracks, leaf_tracks)
        if track:
            product_tracks[source] = track
    for target, source in matched.items():
        if source in product_tracks:
            product_tracks[target] = product_tracks[source]

    legacy_tracks = {
        "生活日用-收纳": "生活日用-收纳用品",
        "其他日用": "生活日用-其他",
        "餐厨水具": "餐厨水具-餐具水具",
        "家居家纺": "家居家纺-家纺",
        "家纺": "家居家纺-家纺",
    }
    for item in items:
        name = text(item.get("name"))
        if name in image_paths:
            item["image"] = image_paths[name]
        track = product_tracks.get(name) or leaf_tracks.get(norm(item.get("leaf")))
        item["category2"] = track or legacy_tracks.get(item.get("category2"), item.get("category2") or "生活日用-其他")

    data["meta"]["imageNote"] = "商品参考图优先使用运营提供的商品图与商品名对照表，本地静态加载"
    SELECTION_JS.write_text(
        "/* 由在线上传+审核后端自动写入 */\n"
        "/* 商品图优先使用运营商品图库；页面按商品名自动匹配本地静态资源 */\n"
        "window.SELECTION_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    payload = {
        "images": image_paths,
        "tracks": product_tracks,
        "leafTracks": leaf_tracks,
        "trackOrder": list(TRACK_NAMES.values()),
    }
    OUT_JS.write_text(
        "/* 由 site/tools/import_product_assets.py 根据运营 Excel 生成 */\n"
        "window.PRODUCT_ASSETS = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )
    print(json.dumps({
        "selectionProducts": len(wanted),
        "matchedNames": len(matched),
        "localImages": len(image_paths),
        "uniqueImageFiles": len(extracted_sources),
        "classifiedProducts": len(product_tracks),
        "assetsMb": round(sum(path.stat().st_size for path in OUT_DIR.glob("*.webp")) / 1024 / 1024, 2),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
