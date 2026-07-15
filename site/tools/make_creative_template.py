#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成【创意卖点·分模块填报模板.xlsx】
- 每个赛道一个 sheet（谁认领谁填），运营只在自己的 sheet 里改，天然不冲突
- 已用现有 creative.js 的数据预填，运营在现成内容上增删改即可
- 顶部有填写说明行，带底色区分「说明区 / 填写区」
用法: python3 tools/make_creative_template.py
产出: 创意卖点填报模板.xlsx（在项目根目录）
"""
import re, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CREATIVE_JS = os.path.join(ROOT, "data", "creative.js")
OUT = os.path.join(os.path.dirname(ROOT), "创意卖点填报模板.xlsx")

# ---------- 从 creative.js 里抠出 JS 对象转成 python ----------
def load_tracks():
    txt = open(CREATIVE_JS, encoding="utf-8").read()
    m = re.search(r"window\.CREATIVE_DATA\s*=\s*(\{.*\});", txt, re.S)
    body = m.group(1)
    # 极简 JS->JSON：给 key 加引号、去注释、去尾逗号
    body = re.sub(r"/\*.*?\*/", "", body, flags=re.S)
    body = re.sub(r"//.*", "", body)
    body = re.sub(r"([{,]\s*)([A-Za-z_]\w*)\s*:", r'\1"\2":', body)
    body = re.sub(r",(\s*[}\]])", r"\1", body)
    return json.loads(body)["tracks"]

# ---------- 样式 ----------
TITLE_FILL = PatternFill("solid", fgColor="2F3A4B")
NOTE_FILL  = PatternFill("solid", fgColor="FFF6E5")
HEAD_FILL  = PatternFill("solid", fgColor="F5B544")
EDIT_FILL  = PatternFill("solid", fgColor="FFFFFF")
white = Font(color="FFFFFF", bold=True, size=13)
head_font = Font(bold=True, size=11)
note_font = Font(color="8A6D1B", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

def style_row(ws, r, ncol, fill, font=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(r, c)
        cell.fill = fill
        if font: cell.font = font
        cell.border = border
        cell.alignment = wrap

def build_track_sheet(wb, t):
    name = t["name"][:28]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = "F5B544"

    # 大标题
    ws.merge_cells("A1:D1")
    ws["A1"] = f"【{t['name']}】赛道 · 卖点创意填报　（负责人：{t.get('owner','待分配')}）"
    ws["A1"].fill = TITLE_FILL; ws["A1"].font = white
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    r = 3
    def section(title, note):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, title); ws.cell(r, 1).fill = HEAD_FILL; ws.cell(r, 1).font = head_font
        ws.cell(r, 1).border = border
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, "填写说明：" + note); style_row(ws, r, 4, NOTE_FILL, note_font)
        r += 1

    # 1) 赛道指标
    section("① 赛道整体指标", "填数字即可；不确定就留原值。creativeCount=本期创意条数")
    m = t["metrics"]
    heads = ["创意条数", "点击率CTR(%)", "3s完播(%)", "转化率CVR(%)", "CPM(元)"]
    for i, h in enumerate(heads):
        ws.cell(r, i + 1, h); ws.cell(r, i + 1).fill = HEAD_FILL; ws.cell(r, i + 1).font = head_font
        ws.cell(r, i + 1).border = border
    r += 1
    vals = [m.get("creativeCount"), m.get("ctr"), m.get("play3s"), m.get("cvr"), m.get("cpm")]
    for i, v in enumerate(vals):
        ws.cell(r, i + 1, v); ws.cell(r, i + 1).fill = EDIT_FILL; ws.cell(r, i + 1).border = border
    r += 2

    # 2) 卖点词云
    section("② 主要卖点词（词云用）", "一行一个词；权重0-100，越大字越大越居中。可增删行")
    ws.cell(r, 1, "卖点词"); ws.cell(r, 2, "权重(0-100)")
    for c in (1, 2):
        ws.cell(r, c).fill = HEAD_FILL; ws.cell(r, c).font = head_font; ws.cell(r, c).border = border
    r += 1
    for w in t.get("sellingWords", []):
        ws.cell(r, 1, w["word"]); ws.cell(r, 2, w.get("weight"))
        ws.cell(r, 1).fill = EDIT_FILL; ws.cell(r, 2).fill = EDIT_FILL
        ws.cell(r, 1).border = border; ws.cell(r, 2).border = border
        r += 1
    r += 1

    # 3) 卖点动因/背景
    section("③ 卖点动因·需求背景（重点！卖点为什么成立）",
            "动因=时令/换季/场景/情绪等背景；需求=由此产生的真实诉求；支撑词=可落到创意的词，用、隔开")
    hs = ["触发动因（背景）", "由此产生的真实需求", "支撑卖点词（用、隔开）"]
    for i, h in enumerate(hs):
        ws.cell(r, i + 1, h); ws.cell(r, i + 1).fill = HEAD_FILL; ws.cell(r, i + 1).font = head_font
        ws.cell(r, i + 1).border = border
    r += 1
    for c in t.get("sellingContext", []):
        ws.cell(r, 1, c.get("driver")); ws.cell(r, 2, c.get("need"))
        ws.cell(r, 3, "、".join(c.get("words", [])))
        for cc in (1, 2, 3):
            ws.cell(r, cc).fill = EDIT_FILL; ws.cell(r, cc).border = border
        r += 1
    r += 1

    # 4) 常见话术
    section("④ 常见爆款话术", "一行一条，直接抄跑量素材里的高转化口播/字幕。可增删行")
    ws.cell(r, 1, "话术"); ws.cell(r, 1).fill = HEAD_FILL; ws.cell(r, 1).font = head_font; ws.cell(r, 1).border = border
    r += 1
    for s in t.get("scripts", []):
        ws.cell(r, 1, s); ws.cell(r, 1).fill = EDIT_FILL; ws.cell(r, 1).border = border
        r += 1
    r += 1

    # 5) 跑量关键点
    section("⑤ 跑量关键点（4个维度）",
            "四个维度：钩子/强对比、节点/季节痛点、福利/紧迫、明星/IP。描述这个赛道怎么打")
    ws.cell(r, 1, "维度"); ws.cell(r, 2, "打法描述")
    for c in (1, 2):
        ws.cell(r, c).fill = HEAD_FILL; ws.cell(r, c).font = head_font; ws.cell(r, c).border = border
    r += 1
    for k in t.get("keyPoints", []):
        ws.cell(r, 1, k.get("title")); ws.cell(r, 2, k.get("desc"))
        ws.cell(r, 1).fill = EDIT_FILL; ws.cell(r, 2).fill = EDIT_FILL
        ws.cell(r, 1).border = border; ws.cell(r, 2).border = border
        r += 1

    # 列宽
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.freeze_panes = "A3"
    return ws

def build_readme_sheet(wb, tracks):
    ws = wb.create_sheet("① 使用说明", 0)
    ws.sheet_properties.tabColor = "37C98A"
    ws.merge_cells("A1:B1")
    ws["A1"] = "创意卖点填报模板 · 使用说明"
    ws["A1"].fill = TITLE_FILL; ws["A1"].font = white
    ws.row_dimensions[1].height = 30
    lines = [
        ("怎么用？", "下方每个赛道一个 sheet（标签页）。找到你负责的赛道，只填你那一页即可。"),
        ("填什么？", "白色格子=可填/可改；黄色=标题；米色=填写说明。在现有内容上增删改就行。"),
        ("能加行吗？", "可以。卖点词/话术/动因不够就直接插入行继续写，不用担心格式。"),
        ("填完怎么办？", "存盘，把这个 Excel 发回给维护同学（葱），一键就能更新到网页，你不用碰代码。"),
        ("多久更新？", "建议每周或每次发布前统一收一次表，维护同学同步后网页即为最新对客版本。"),
        ("负责人分工", "、".join(f"{t['name']}={t.get('owner','待分配')}" for t in tracks)),
    ]
    r = 3
    for k, v in lines:
        ws.cell(r, 1, k); ws.cell(r, 1).fill = HEAD_FILL; ws.cell(r, 1).font = head_font; ws.cell(r, 1).border = border
        ws.cell(r, 2, v); ws.cell(r, 2).fill = NOTE_FILL; ws.cell(r, 2).alignment = wrap; ws.cell(r, 2).border = border
        ws.row_dimensions[r].height = 34
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80
    return ws

def main():
    tracks = load_tracks()
    wb = Workbook()
    wb.remove(wb.active)
    build_readme_sheet(wb, tracks)
    for t in tracks:
        build_track_sheet(wb, t)
    wb.save(OUT)
    print("已生成填报模板:", OUT)
    print("赛道 sheet:", [t["name"] for t in tracks])

if __name__ == "__main__":
    main()
