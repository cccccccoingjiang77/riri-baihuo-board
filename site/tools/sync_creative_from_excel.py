#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
【一键同步】把运营填好的《创意卖点填报模板.xlsx》回写进 data/creative.js
策略：以现有 creative.js 为基底，只用 Excel 中填写的字段覆盖，
      未在表格维护的字段（Top素材/关键帧/videoUrl 等）原样保留，绝不误删。

用法:
    python3 tools/sync_creative_from_excel.py [可选:xlsx路径 或 文件夹路径]
    - 不带参数：默认读项目根目录的《创意卖点填报模板.xlsx》（总模板，多sheet）
    - 传一个 xlsx 文件：读该总模板
    - 传一个文件夹：读该文件夹下所有 xlsx（分赛道模式，每份一个赛道）
      若不带参数且存在《填报-分赛道》目录，也会自动合并该目录下的分表
可选一键推送:
    python3 tools/sync_creative_from_excel.py --push
"""
import re, json, os, sys, subprocess, datetime
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))       # site/
PROJ = os.path.dirname(ROOT)                                            # 项目根
CREATIVE_JS = os.path.join(ROOT, "data", "creative.js")
DEFAULT_XLSX = os.path.join(PROJ, "创意卖点填报模板.xlsx")
SPLIT_DIR = os.path.join(PROJ, "填报-分赛道")

# ---------- 读现有 creative.js（作为基底，保留未维护字段） ----------
def load_current():
    txt = open(CREATIVE_JS, encoding="utf-8").read()
    m = re.search(r"window\.CREATIVE_DATA\s*=\s*(\{.*\});", txt, re.S)
    body = m.group(1)
    body = re.sub(r"/\*.*?\*/", "", body, flags=re.S)
    body = re.sub(r"//.*", "", body)
    body = re.sub(r"([{,]\s*)([A-Za-z_]\w*)\s*:", r'\1"\2":', body)
    body = re.sub(r",(\s*[}\]])", r"\1", body)
    return json.loads(body)

def cell(ws, r, c):
    v = ws.cell(r, c).value
    return "" if v is None else v

def to_num(v):
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except:
        return v if v not in ("", None) else None

# ---------- 从单个赛道 sheet 抽取运营填写的内容 ----------
def parse_track_sheet(ws):
    """按模板固定小节顺序解析：① 指标 ② 卖点词 ③ 动因 ④ 话术 ⑤ 关键点"""
    out = {}
    maxr = ws.max_row
    # 找每个小节的表头行（靠关键字定位，容忍运营插入/删除行）
    # exact=True 时要求单元格值完全等于 keyword，避免匹配到说明行
    def find(keyword, col=1, exact=False):
        for r in range(1, maxr + 1):
            v = str(cell(ws, r, col)).strip()
            if (v == keyword) if exact else (keyword in v):
                return r
        return None

    # ① 指标：表头下一行取5个数（用 exact 精确匹配表头，避免命中含"创意条数"的说明行）
    r = find("创意条数", exact=True)
    if r:
        out["metrics"] = {
            "creativeCount": to_num(cell(ws, r + 1, 1)),
            "ctr": to_num(cell(ws, r + 1, 2)),
            "play3s": to_num(cell(ws, r + 1, 3)),
            "cvr": to_num(cell(ws, r + 1, 4)),
            "cpm": to_num(cell(ws, r + 1, 5)),
        }

    # ② 卖点词：从"卖点词/权重"表头下往下读，直到空行
    r = find("卖点词", exact=True)
    words = []
    if r:
        i = r + 1
        while i <= maxr and str(cell(ws, i, 1)).strip():
            w = str(cell(ws, i, 1)).strip()
            wt = to_num(cell(ws, i, 2)) or 0
            words.append({"word": w, "weight": wt})
            i += 1
    if words:
        out["sellingWords"] = words

    # ③ 动因：三列
    r = find("触发动因（背景）")
    ctx = []
    if r:
        i = r + 1
        while i <= maxr and str(cell(ws, i, 1)).strip():
            drv = str(cell(ws, i, 1)).strip()
            need = str(cell(ws, i, 2)).strip()
            ws_words = [x.strip() for x in re.split(r"[、,，]", str(cell(ws, i, 3))) if x.strip()]
            ctx.append({"driver": drv, "need": need, "words": ws_words})
            i += 1
    if ctx:
        out["sellingContext"] = ctx

    # ④ 话术：单列
    r = find("话术", exact=True)
    scripts = []
    if r:
        i = r + 1
        while i <= maxr and str(cell(ws, i, 1)).strip():
            t = str(cell(ws, i, 1)).strip()
            # 遇到下一小节标题（含"维度"/"跑量"）就停
            if "维度" in t or "跑量" in t:
                break
            scripts.append(t)
            i += 1
    if scripts:
        out["scripts"] = scripts

    # ⑤ 关键点：维度 + 描述
    r = find("维度", exact=True)
    kps = []
    if r:
        i = r + 1
        while i <= maxr and str(cell(ws, i, 1)).strip():
            kps.append({"title": str(cell(ws, i, 1)).strip(),
                        "desc": str(cell(ws, i, 2)).strip()})
            i += 1
    if kps:
        out["keyPoints"] = kps
    return out

def dump_js(data):
    lines = ["/* 由 sync_creative_from_excel.py 从《创意卖点填报模板.xlsx》同步生成 */",
             "/* 运营在 Excel 里填，本文件自动生成，勿手改本文件 */",
             "window.CREATIVE_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";",
             ""]
    return "\n".join(lines)

def collect_xlsx(path):
    """把输入路径归一化为一批 xlsx 文件列表。"""
    if os.path.isdir(path):
        files = [os.path.join(path, f) for f in sorted(os.listdir(path))
                 if f.lower().endswith(".xlsx") and not f.startswith("~$")]
        return files
    return [path]

def apply_workbook(xlsx, name2track):
    """读一个 xlsx，把其中的赛道 sheet 合并进 name2track，返回本文件更新的赛道名。"""
    updated = []
    wb = load_workbook(xlsx, data_only=True)
    for sh in wb.sheetnames:
        if sh not in name2track:      # 跳过"使用说明"等非赛道页
            continue
        parsed = parse_track_sheet(wb[sh])
        if parsed:
            name2track[sh].update(parsed)   # 只覆盖填写字段，其余（Top素材等）保留
            updated.append(sh)
    return updated

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    do_push = "--push" in sys.argv

    # 决定数据源：显式路径 > 分赛道目录 > 总模板
    if args:
        src = args[0]
    elif os.path.isdir(SPLIT_DIR) and any(
            f.lower().endswith(".xlsx") and not f.startswith("~$")
            for f in os.listdir(SPLIT_DIR)):
        src = SPLIT_DIR
    else:
        src = DEFAULT_XLSX

    files = collect_xlsx(src)
    files = [f for f in files if os.path.exists(f)]
    if not files:
        print("找不到填报表:", src); sys.exit(1)

    data = load_current()
    name2track = {t["name"]: t for t in data["tracks"]}

    updated = []
    for f in files:
        got = apply_workbook(f, name2track)
        if got:
            print(f"  读入 {os.path.basename(f)} → {got}")
        updated.extend(got)

    if not updated:
        print("没有识别到任何赛道数据（检查 sheet 名是否与赛道名一致）。")

    data["meta"]["updatedAt"] = datetime.date.today().isoformat()
    open(CREATIVE_JS, "w", encoding="utf-8").write(dump_js(data))
    print("已同步赛道:", sorted(set(updated)))
    print("写入:", CREATIVE_JS)

    if do_push:
        try:
            subprocess.run(["git", "-C", PROJ, "add", "-A"], check=True)
            subprocess.run(["git", "-C", PROJ,
                            "-c", "user.email=cccccccoingjiang77@users.noreply.github.com",
                            "-c", "user.name=cccccccoingjiang77",
                            "commit", "-m", f"chore: 同步创意卖点数据（{datetime.date.today()}）"], check=True)
            subprocess.run(["git", "-C", PROJ, "push"], check=True)
            print("已推送到 GitHub。")
        except subprocess.CalledProcessError as e:
            print("推送失败（可能无变更或需检查网络）:", e)

if __name__ == "__main__":
    main()
