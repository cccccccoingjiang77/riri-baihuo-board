#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把「一个赛道的结构化 JSON」生成一份标准的《填报-XX赛道.xlsx》。
产出的表格排版 / sheet 名 / 表头关键字，与维护同学的回收脚本
sync_creative_from_excel.py 完全对齐——赛道同学把它发回给葱即可一键同步。

用法:
    python3 scripts/build_track_excel.py input.json [可选:输出目录]
    python3 scripts/build_track_excel.py scripts/example-清洁工具.json   # 用黄金样例跑通

input.json 结构见 references/schema.md 末尾。
"""
import sys, os, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------- 样式（与总模板一致，确保回收脚本能识别） ----------
TITLE_FILL = PatternFill("solid", fgColor="2F3A4B")
NOTE_FILL  = PatternFill("solid", fgColor="FFF6E5")
HEAD_FILL  = PatternFill("solid", fgColor="F5B544")
EDIT_FILL  = PatternFill("solid", fgColor="FFFFFF")
white = Font(color="FFFFFF", bold=True, size=13)
head_font = Font(bold=True, size=11)
note_font = Font(color="8A6D1B", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

VALID_TRACKS = {"家纺", "生活日用-清洁工具", "生活日用-功效品",
                "生活日用-收纳", "餐厨水具"}


def style_row(ws, r, ncol, fill, font=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(r, c)
        cell.fill = fill
        if font:
            cell.font = font
        cell.border = border
        cell.alignment = wrap


def build(ws, t):
    ws.sheet_properties.tabColor = "F5B544"
    # 大标题
    ws.merge_cells("A1:D1")
    ws["A1"] = f"【{t['name']}】赛道 · 卖点创意填报　（负责人：{t.get('owner', '待分配')}）"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = white
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = "★ 本表由『赛道卖点创意分析 skill』生成，格式已对齐回收脚本。直接发回给维护同学（葱）即可一键同步进网页。"
    style_row(ws, 2, 4, NOTE_FILL, note_font)
    ws.row_dimensions[2].height = 22

    r = 4

    def section(title, note):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, title)
        ws.cell(r, 1).fill = HEAD_FILL
        ws.cell(r, 1).font = head_font
        ws.cell(r, 1).border = border
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, "填写说明：" + note)
        style_row(ws, r, 4, NOTE_FILL, note_font)
        r += 1

    # ① 指标
    section("① 赛道整体指标", "填数字即可；不确定就留原值。creativeCount=本期创意条数")
    m = t.get("metrics", {})
    heads = ["创意条数", "点击率CTR(%)", "3s完播(%)", "转化率CVR(%)", "CPM(元)"]
    for i, h in enumerate(heads):
        ws.cell(r, i + 1, h)
        ws.cell(r, i + 1).fill = HEAD_FILL
        ws.cell(r, i + 1).font = head_font
        ws.cell(r, i + 1).border = border
    r += 1
    vals = [m.get("creativeCount"), m.get("ctr"), m.get("play3s"), m.get("cvr"), m.get("cpm")]
    for i, v in enumerate(vals):
        ws.cell(r, i + 1, v)
        ws.cell(r, i + 1).fill = EDIT_FILL
        ws.cell(r, i + 1).border = border
    r += 2

    # ② 卖点词
    section("② 主要卖点词（词云用）", "一行一个词；权重0-100，越大字越大越居中。可增删行")
    ws.cell(r, 1, "卖点词")
    ws.cell(r, 2, "权重(0-100)")
    for c in (1, 2):
        ws.cell(r, c).fill = HEAD_FILL
        ws.cell(r, c).font = head_font
        ws.cell(r, c).border = border
    r += 1
    for w in t.get("sellingWords", []):
        ws.cell(r, 1, w.get("word"))
        ws.cell(r, 2, w.get("weight"))
        ws.cell(r, 1).fill = EDIT_FILL
        ws.cell(r, 2).fill = EDIT_FILL
        ws.cell(r, 1).border = border
        ws.cell(r, 2).border = border
        r += 1
    r += 1

    # ③ 动因
    section("③ 卖点动因·需求背景（重点！卖点为什么成立）",
            "动因=时令/换季/场景/情绪等背景；需求=由此产生的真实诉求；支撑词=可落到创意的词，用、隔开")
    hs = ["触发动因（背景）", "由此产生的真实需求", "支撑卖点词（用、隔开）"]
    for i, h in enumerate(hs):
        ws.cell(r, i + 1, h)
        ws.cell(r, i + 1).fill = HEAD_FILL
        ws.cell(r, i + 1).font = head_font
        ws.cell(r, i + 1).border = border
    r += 1
    for c in t.get("sellingContext", []):
        ws.cell(r, 1, c.get("driver"))
        ws.cell(r, 2, c.get("need"))
        ws.cell(r, 3, "、".join(c.get("words", [])))
        for cc in (1, 2, 3):
            ws.cell(r, cc).fill = EDIT_FILL
            ws.cell(r, cc).border = border
        r += 1
    r += 1

    # ④ 话术
    section("④ 常见爆款话术", "一行一条，直接抄跑量素材里的高转化口播/字幕。可增删行")
    ws.cell(r, 1, "话术")
    ws.cell(r, 1).fill = HEAD_FILL
    ws.cell(r, 1).font = head_font
    ws.cell(r, 1).border = border
    r += 1
    for s in t.get("scripts", []):
        ws.cell(r, 1, s)
        ws.cell(r, 1).fill = EDIT_FILL
        ws.cell(r, 1).border = border
        r += 1
    r += 1

    # ⑤ 关键点
    section("⑤ 跑量关键点（4个维度）",
            "四个维度：钩子/强对比、节点/季节痛点、福利/紧迫、明星/IP。描述这个赛道怎么打")
    ws.cell(r, 1, "维度")
    ws.cell(r, 2, "打法描述")
    for c in (1, 2):
        ws.cell(r, c).fill = HEAD_FILL
        ws.cell(r, c).font = head_font
        ws.cell(r, c).border = border
    r += 1
    for k in t.get("keyPoints", []):
        ws.cell(r, 1, k.get("title"))
        ws.cell(r, 2, k.get("desc"))
        ws.cell(r, 1).fill = EDIT_FILL
        ws.cell(r, 2).fill = EDIT_FILL
        ws.cell(r, 1).border = border
        ws.cell(r, 2).border = border
        r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.freeze_panes = "A4"


def safe_name(name):
    return re.sub(r'[\\/:*?"<>|]', "_", name)


def main():
    if len(sys.argv) < 2:
        print("用法: python3 scripts/build_track_excel.py input.json [输出目录]")
        sys.exit(1)
    inp = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else os.getcwd()

    t = json.load(open(inp, encoding="utf-8"))
    name = t.get("name", "").strip()
    if not name:
        print("input.json 缺少 name 字段（赛道名）。")
        sys.exit(1)
    if name not in VALID_TRACKS:
        print(f"⚠ 警告：赛道名『{name}』不在标准列表 {sorted(VALID_TRACKS)} 中，"
              "回收时可能无法识别（除非网页里已有同名赛道）。")

    wb = Workbook()
    ws = wb.active
    ws.title = name[:28]        # sheet 名 = 赛道名，供回收脚本识别
    build(ws, t)

    os.makedirs(out_dir, exist_ok=True)
    fname = f"填报-{safe_name(name)}.xlsx"
    out_path = os.path.join(out_dir, fname)
    wb.save(out_path)
    print("已生成:", out_path)
    print("sheet 名:", ws.title, "（= 赛道名，回收脚本据此识别）")
    print("直接把这个文件发回给维护同学即可。")


if __name__ == "__main__":
    main()
